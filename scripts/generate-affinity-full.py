#!/usr/bin/env python3
"""从 data/AI 亲和原则全量.xlsx 生成 principles-affinity-full.html（不含 D/F 列）。

变更：
- 按四节点分组（找的到 / 找的准 / 读的懂 / 读的顺），每组独立表格
- 原则名称列下方显示来源徽标（实测诊断 / 根因泛化 / 行业标准 / 待实测）
"""
import html
import re
import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "data" / "AI 亲和原则全量.xlsx"
OUTPUT = ROOT / "principles-affinity-full.html"

NUM_PREFIX = re.compile(r"^\d+\.\s*")
LETTER_PREFIX = re.compile(r"^[a-zA-Z]\.\s*")

# ── 四节点映射（Excel 序号 → 节点） ──────────────────────────────────────────
NODE_MAP: dict[int, str] = {
    # 找的到（3条）
    1: "找的到", 26: "找的到", 29: "找的到",
    # 找的准（7条）
    8: "找的准", 10: "找的准", 11: "找的准", 12: "找的准",
    27: "找的准", 31: "找的准", 33: "找的准",
    # 读的懂（13条）
    5: "读的懂", 9: "读的懂", 15: "读的懂", 17: "读的懂",
    18: "读的懂", 19: "读的懂", 20: "读的懂", 21: "读的懂",
    22: "读的懂", 23: "读的懂", 24: "读的懂", 25: "读的懂", 30: "读的懂",
    # 读的顺（12条）
    2: "读的顺", 3: "读的顺", 4: "读的顺", 6: "读的顺", 7: "读的顺",
    13: "读的顺", 14: "读的顺", 16: "读的顺", 28: "读的顺",
    32: "读的顺", 34: "读的顺", 35: "读的顺",
}

NODE_ORDER = ["找的到", "找的准", "读的懂", "读的顺"]

NODE_DESC: dict[str, str] = {
    "找的到": "AI 可发现文档入口、索引文档目录",
    "找的准": "AI 可定位到正确版本与精准范围",
    "读的懂": "AI 可解析、提取结构化内容",
    "读的顺": "AI 可理解语义、应用于生成与推理",
}

# ── 来源映射（Excel 序号 → 来源类型） ─────────────────────────────────────────
SOURCE_MAP: dict[int, str] = {
    # 实测原则（13条）
    1: "实测诊断", 10: "实测诊断", 11: "实测诊断",
    17: "实测诊断", 18: "实测诊断", 20: "实测诊断", 22: "实测诊断",
    23: "实测诊断", 24: "实测诊断", 25: "实测诊断", 26: "实测诊断", 27: "实测诊断", 35: "实测诊断",
    # 根因泛化（16条）
    2: "根因泛化", 3: "根因泛化", 5: "根因泛化", 7: "根因泛化", 9: "根因泛化",
    12: "根因泛化", 13: "根因泛化", 15: "根因泛化", 16: "根因泛化", 19: "根因泛化",
    21: "根因泛化", 28: "根因泛化", 29: "根因泛化", 31: "根因泛化", 34: "根因泛化",
    # 行业标准（6条）
    4: "行业标准", 6: "行业标准", 8: "行业标准", 30: "行业标准",
    32: "行业标准", 33: "行业标准",
}

SOURCE_BADGE: dict[str, str] = {
    "实测诊断": "badge-ok",
    "根因泛化": "badge-should",
    "行业标准": "badge-primary",
    "待实测": "badge-muted",
}


# ── 文本格式化 ────────────────────────────────────────────────────────────────

def is_cell_title_line(line: str) -> bool:
    if NUM_PREFIX.match(line) or LETTER_PREFIX.match(line):
        return False
    return "：" in line or ":" in line


def fmt_line(line: str) -> str:
    line = line.strip()
    if not line:
        return ""
    if not is_cell_title_line(line):
        return html.escape(line)
    for sep in ("：", ":"):
        if sep in line:
            head, tail = line.split(sep, 1)
            head = head.strip()
            if head:
                return f"<strong>{html.escape(head)}</strong>{sep}{html.escape(tail)}"
            break
    return html.escape(line)


def fmt_cell(text) -> str:
    if text is None:
        return "—"
    s = str(text).strip()
    if not s or s in ("-", "/"):
        return "—"
    return "<br>".join(fmt_line(line) for line in s.split("\n") if line.strip())


def has_adjustment(text) -> bool:
    if text is None:
        return False
    s = str(text).strip()
    return bool(s) and s not in ("-", "/", "—")


# ── 数据读取 ──────────────────────────────────────────────────────────────────

def build_rows(xlsx: Path) -> list[tuple]:
    import openpyxl
    ws = openpyxl.load_workbook(xlsx, data_only=True).active
    rows = []
    for r in range(2, ws.max_row + 1):
        seq = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        if seq is None or not name:
            continue
        content = ws.cell(r, 3).value
        design = ws.cell(r, 5).value
        dev = ws.cell(r, 7).value
        rows.append((int(seq), str(name).strip(), content, design, dev))
    return rows


# ── HTML 构建 ─────────────────────────────────────────────────────────────────

TABLE_TEMPLATE = """\
      <div class="affinity-table-wrap">
      <table class="data-table data-table--affinity">
        <colgroup>
          <col class="col-id">
          <col class="col-item">
          <col class="col-content">
          <col class="col-design">
          <col class="col-code">
          <col class="col-source">
        </colgroup>
        <thead>
          <tr>
            <th>序号</th>
            <th>原则名称</th>
            <th>文档内容调整</th>
            <th>设计UI 调整</th>
            <th>前端调整</th>
            <th>原则来源</th>
          </tr>
        </thead>
        <tbody>
{tbody}
        </tbody>
      </table>
      </div>"""


PROBLEM_MAP: dict[int, str] = {
    1:  "problems-structure-llms.html",
    10: "problems-timeliness.html",
    11: "problems-timeliness.html",
    17: "problems-content-image.html",
    18: "problems-content-hotzone.html",
    20: "problems-content-link.html",
    22: "problems-content-code.html",
    23: "problems-content-table.html",
    24: "problems-content-note.html",
    25: "problems-content-tab.html",
    27: "problems-structure-metadata.html",
    35: "problems-format.html",
}


def build_row_html(seq: int, name: str, content, design, dev) -> str:
    row_id = f"principle-full-{seq}"
    source = SOURCE_MAP.get(seq, "根因泛化")
    badge_cls = SOURCE_BADGE[source]
    attrs = (
        f'data-adjust-content="{"1" if has_adjustment(content) else "0"}"'
        f' data-adjust-design="{"1" if has_adjustment(design) else "0"}"'
        f' data-adjust-dev="{"1" if has_adjustment(dev) else "0"}"'
    )
    problem_href = PROBLEM_MAP.get(seq)
    if source == "实测诊断":
        badge = (
            f'<a href="{problem_href}" class="badge badge-accent source-diagnosis-link">实测诊断</a>'
            if problem_href
            else '<span class="badge badge-accent">实测诊断</span>'
        )
        source_td = f'            <td class="col-source">{badge}</td>'
    else:
        source_td = '            <td class="col-source">—</td>'
    return "\n".join([
        f'          <tr id="{row_id}" {attrs}>',
        f'            <td class="col-id">{seq}</td>',
        f'            <td>{html.escape(name)}</td>',
        f"            <td>{fmt_cell(content)}</td>",
        f"            <td>{fmt_cell(design)}</td>",
        f"            <td>{fmt_cell(dev)}</td>",
        source_td,
        "          </tr>",
    ])


def build_grouped_sections(rows: list[tuple]) -> str:
    # 按节点分组，保持 Excel 序号顺序
    groups: dict[str, list[tuple]] = defaultdict(list)
    for row in rows:
        seq = row[0]
        node = NODE_MAP.get(seq, "读的顺")
        groups[node].append(row)

    parts = []
    for node in NODE_ORDER:
        node_rows = groups.get(node, [])
        if not node_rows:
            continue
        tbody_lines = [build_row_html(*r) for r in node_rows]
        tbody = "\n".join(tbody_lines)
        table_html = TABLE_TEMPLATE.format(tbody=tbody)
        desc = NODE_DESC.get(node, "")
        count = len(node_rows)
        parts.append(
            f'      <h3 class="affinity-group-title">{html.escape(node)}</h3>\n'
            + table_html
        )
    return "\n\n".join(parts)


def build_html(grouped_sections: str, total: int) -> str:
    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>全量亲和原则 · 亲和原则 · 社区AI亲和分析</title>
  <link rel="stylesheet" href="assets/css/style.css">
</head>
<body class="inner-page" data-module="principles" data-page="affinity-full">

<header class="site-header">
  <div class="site-logo">社区AI亲和分析</div>
  <nav class="site-nav">
    <a href="index.html">首页</a>
    <a href="background-motivation.html">研究概览</a>
    <a href="problems-answer-search.html">问题论证</a>
    <a href="principles-affinity.html" class="active">亲和原则</a>
  </nav>
</header>

<div class="page-wrapper">

  <aside id="module-sidebar"></aside>

  <main class="main-content">

    <div class="page-header">
      <h1>全量亲和原则 35 项</h1>
      <p class="page-desc">按「找的到 / 找的准 / 读的懂 / 读的顺」四节点框架分组，覆盖文档内容、设计 UI、前端三个交付维度。默认展示实测诊断原则。</p>
    </div>

    <div class="affinity-full-toolbar">
      <h2 class="affinity-full-summary" id="affinity-full-summary" aria-live="polite">全部修改项 {total} 项</h2>
      <div class="affinity-full-filter">
        <select id="affinity-full-filter" class="affinity-full-filter-select" aria-label="调整项筛选">
          <option value="all" selected>调整项：全部</option>
          <option value="content">调整项：文档内容</option>
          <option value="design">调整项：设计UI</option>
          <option value="dev">调整项：前端</option>
        </select>
      </div>
    </div>

    <section class="section" id="affinity-catalog">

{grouped_sections}

    </section>

  </main></div>

<footer class="site-footer">
  模块 03 · 亲和原则 · 社区AI亲和分析
</footer>

<script src="assets/js/site-config.js"></script>
<script src="assets/js/site-init.js"></script>
<script src="assets/js/module-sidebar.js"></script>
<script src="assets/js/affinity-full-filter.js"></script>
</body>
</html>
"""


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.is_file():
        sys.exit(f"找不到 Excel：{xlsx}")
    rows = build_rows(xlsx)
    grouped = build_grouped_sections(rows)
    OUTPUT.write_text(build_html(grouped, len(rows)), encoding="utf-8")
    print(f"已生成 {OUTPUT}（{len(rows)} 行，四节点分组）")


if __name__ == "__main__":
    main()
