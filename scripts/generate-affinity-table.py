#!/usr/bin/env python3
"""从 亲和原则汇总 NEW.xlsx 生成 principles-affinity.html 表格 tbody。"""
import html
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path.home() / "Downloads" / "亲和原则汇总NEW.xlsx"
HTML = ROOT / "principles-affinity.html"
TBODY_SNIPPET = ROOT / "scripts" / "_affinity_tbody.html"
NUM_PREFIX = re.compile(r"^\d+\.\s*")
LETTER_PREFIX = re.compile(r"^[a-zA-Z]\.\s*")


def is_cell_title_line(line: str) -> bool:
    """单元格内标题：非数字/字母编号条目，且以冒号分段。"""
    if NUM_PREFIX.match(line) or LETTER_PREFIX.match(line):
        return False
    return "：" in line or ":" in line


def fmt_line(line: str) -> str:
    line = line.strip()
    if not line:
        return ""
    if not is_cell_title_line(line):
        return html.escape(line)
    for sep in ("：", ":"):
        if sep in line:
            head, tail = line.split(sep, 1)
            head = head.strip()
            if head:
                return f"<strong>{html.escape(head)}</strong>{sep}{html.escape(tail)}"
            break
    return html.escape(line)


def fmt_cell(text):
    if text is None:
        return "—"
    s = str(text).strip()
    if not s or s == "-":
        return "—"
    return "<br>".join(fmt_line(line) for line in s.split("\n") if line.strip())


def norm_code(raw) -> str:
    if raw is None:
        return "—"
    return str(raw).strip().upper()


def strip_domain_prefix(domain: str) -> str:
    s = str(domain).strip()
    return re.sub(r"^[A-D]\.\s*", "", s)


def build_tbody(xlsx: Path) -> str:
    import openpyxl

    ws = openpyxl.load_workbook(xlsx, data_only=True).active
    rows = []
    current_domain = ""
    for r in range(2, ws.max_row + 1):
        domain = ws.cell(r, 1).value
        code = ws.cell(r, 2).value
        name = ws.cell(r, 3).value
        if not code and not name:
            continue
        if domain:
            current_domain = str(domain).strip()
        design = ws.cell(r, 4).value
        content = ws.cell(r, 5).value
        dev = ws.cell(r, 6).value
        rows.append((current_domain, norm_code(code), str(name or "").strip(), design, content, dev))

    lines = []
    seq = 0
    i = 0
    while i < len(rows):
        domain = rows[i][0]
        j = i + 1
        while j < len(rows) and rows[j][0] == domain:
            j += 1
        span = j - i
        dim_label = strip_domain_prefix(domain)
        for k in range(span):
            domain, code, name, design, content, dev = rows[i + k]
            seq += 1
            row_id = f' id="principle-{code.lower()}"' if re.match(r"^[A-D]\d+$", code) else ""
            parts = [f"          <tr{row_id}>"]
            if k == 0:
                parts.append(
                    f'            <td class="col-dim" rowspan="{span}">{html.escape(dim_label)}</td>'
                )
            parts.extend(
                [
                    f'            <td class="col-id">{seq}</td>',
                    f"            <td>{html.escape(name)}</td>",
                    f"            <td>{fmt_cell(design)}</td>",
                    f"            <td>{fmt_cell(content)}</td>",
                    f"            <td>{fmt_cell(dev)}</td>",
                    "          </tr>",
                ]
            )
            lines.append("\n".join(parts))
        i = j
    return "\n".join(lines)


def patch_html(tbody: str) -> None:
    text = HTML.read_text(encoding="utf-8")
    text = re.sub(
        r"<tbody>.*?</tbody>",
        "<tbody>\n" + tbody + "\n        </tbody>",
        text,
        count=1,
        flags=re.S,
    )
    HTML.write_text(text, encoding="utf-8")


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.is_file():
        sys.exit(f"找不到 Excel：{xlsx}")
    tbody = build_tbody(xlsx)
    TBODY_SNIPPET.write_text(tbody, encoding="utf-8")
    patch_html(tbody)
    print(f"已更新 {HTML}（{tbody.count('<tr')} 行）")


if __name__ == "__main__":
    main()
